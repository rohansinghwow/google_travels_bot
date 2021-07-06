import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

PATH = "chromedriver.exe" # Chrome Driver location
driver = webdriver.Chrome(PATH)



# tourist_titles = []
# tourist_desc = []
# tourist_ratings = []


def excel_stuff(): #Function to get data one by one from excel file named as Europe.xlsx

        wb = openpyxl.load_workbook('Europe.xlsx')
        sheets = wb.sheetnames
        sh1 = wb['Sheet1']
        row = sh1.max_row
        places_names = []
        for i in range(3, row+1):
            places_names.append((sh1.cell(i, 1).value))




        for places_single in places_names: # Take each row and get the City name from Excel file and put it into places_single variable


            driver.get("https://www.google.com/travel/?dest_src=ut&tcfs=UgJgAQ&ved=2ahUKEwigxuvT-KLxAhVZhmYCHdjnDrIQyJABegQIABAR&ictx=2#ttdm=48.193695_16.350409_13") #The main URL of google travels where the search bar appears.
            search = driver.find_element_by_class_name('YMlIz')  #search the HTML Page for a div with class name 'YMlIz' which is the search input
            search.send_keys(places_single)  #Input the  single value taken from Excel sheet row
            search.send_keys(Keys.RETURN)  #Press Enter to search
            get_data()  # Run the get_data() function
            driver.back()  # After get_data() finishes then go back to main search page
            time.sleep(1.1)  # Sleep for 1.1 seconds


def get_data(): # Function to put data in search feild one by one and extract the required info into tourist_title[] , toursit_desc[] and tourist_ratings[] list

    container = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CLASS_NAME, "kQb6Eb"))) #Get the div with class name - kQb6Eb which is the contains all the elements as childs
    card_outer = container.find_elements_by_class_name('f4hh3d') #Get all the list of elements with div name  f4hh3d
    for el in card_outer:  # GEt each element from card_outer and name it el
        el.click()  # click a single element from the card_outer list

        placeName = WebDriverWait(driver, 200).until(EC.presence_of_element_located((By.CLASS_NAME, 'enpmuc'))) #Bot waits for 200 seconds to check if the element with class name 'enpmuc' has apperead
        # tourist_titles.append(placeName.text) #Get the text value and Add values to toursit_titles list
        print(placeName.text)

        ratings = WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.CLASS_NAME, 'BgYkof')))
        # tourist_ratings.append(ratings) #Get the text value and Add values to toursit_ratings list
        print(ratings.text)

        desc = WebDriverWait(driver, 400).until(EC.presence_of_element_located((By.CLASS_NAME, 'NYYuTb')))
        # tourist_desc.append(placeName.text) #Get the text value and Add values to tourist_desc list
        print(desc.text)

        close_btn = WebDriverWait(driver, 500).until(EC.presence_of_element_located((By.CLASS_NAME, 'tPsbO'))) #Get cards close button by class name
        close_btn.click()  # CLick the close button
        time.sleep(0.50)  # pause for 0.50 seconds








try:
    excel_stuff()  # Run excel_stuff() function

except:

    driver.quit()   #Close bot and exit




